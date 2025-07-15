import csv
import io
import re
import msoffcrypto

from typing import Optional

import pdfplumber
from pypdf import PdfWriter
from backend.services import csvTextProcessor, pdfTextProcesser
from backend.services.processExported import processExportedExcel
from backend.utils import pdfReader
from backend.utils.detectBank import detectBank
from backend.utils.pdfReader import convertToPypdf, decryptPdf
from openpyxl import load_workbook

import traceback

def fileParser(content: bytes, extension: str, password: Optional[str]):
    if (extension == 'pdf'):
        return parsePdf(content, password)
    elif (extension == 'xlsx'):
        return parseXlsx(content, password)
    elif (extension == 'txt'):
        return parseTxt(content)
    """ elif (extension == 'csv'):
        return parseCSV(content) """

def parseXlsx(content: bytes, password: Optional[str]):
    stream = io.BytesIO(content)
    officeFile = msoffcrypto.OfficeFile(stream)
    if (officeFile.is_encrypted()):
        if password == None:
            return (False, 'Require Password')
        else:
            try:
                decryptedSteam = io.BytesIO()
                officeFile.load_key(password)
                officeFile.decrypt(decryptedSteam)
                stream = decryptedSteam
            except:
                return (False, 'Invalid Password') 
    
    workbook = load_workbook(stream, data_only=True)
    sheetnames = workbook.sheetnames
    if (sheetnames.index('Accounts') != -1 and sheetnames.index('Transactions') != -1):
        return processExportedExcel(workbook)
    else:   
        return (False, 'Please use exported excel file only')

def parsePdf(content: bytes, password: Optional[str]):
    pypdf = convertToPypdf(content)
    if (pdfReader.isPasswordProtected(pypdf)):
        if password == None:
            return (False, 'Require Password')
        else:
            pypdf = decryptPdf(pypdf, password)
            if pypdf == None:
                return (False, 'Invalid Password')
    
    #Remove copy protection (Encryption != password only)
    pypdf = pdfReader.stripEncryption(pypdf)
    extractedText = pdfReader.extractPdfText(pypdf)

    if extractedText == []:
        return (False, 'File cannot be read, ensure it is text and can be selected')
    else:
        bank = detectBank(extractedText)
        return extractTableDetailsFromText(bank, extractedText, pypdf)

#DBS is shit
def parseCSV(content: bytes):
    csv_text_stream = io.StringIO(content.decode('utf-8'))
    reader = csv.reader(csv_text_stream)

    try:    
        return csvTextProcessor.processDBS(reader)
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        last_call = tb[-1]# Only the last call
        print(f"Error in file {last_call.filename}, line {last_call.lineno}, in {last_call.name}\nMessage: {str(e)}")
        return (False, 'Invalid bank type. Please use supported bank types only.')

def parseTxt(content: bytes):
    decodedString = content.decode('utf-8')
    split = decodedString.split('\n')
    extractedText = []
    for each in split:
        stripped = re.sub(r'[^\x20-\x7E\n\r\t]', '', each.strip())
        if len(stripped) > 0:
            extractedText.append(stripped)
    bank = pdfTextProcesser.detectBank(extractedText)
    return extractTableDetailsFromText(bank, extractedText)

def extractTableDetailsFromText(bank, extractedText, pypdf):
    try:
        match bank:
            case 'DBS':
                return pdfTextProcesser.processDBS(extractedText)
            case 'UOB':
                return pdfTextProcesser.processUOB(extractedText)
            case 'OCBC':
                # pdf have weird formatting or some shit for ocbc pdf
                # takes very long for pdf plumber to extract text 
                # similar object count in each page to other pdf
                # see objects = page.objects
                # might be char object got some stupid complicated formatting shit
                    
                return pdfTextProcesser.processOCBC(extractedText)     
            case 'SC':
                return pdfTextProcesser.processSC(extractedText)
            case _:
                return pdfTextProcesser.processOthers(extractedText, pypdf)
            
    except Exception as e:
        # prints relavent part of the error for easier debugging
        tb = traceback.extract_tb(e.__traceback__)
        last_call = tb[-1]# Only the last call
        print(f"Error in file {last_call.filename}, line {last_call.lineno}, in {last_call.name}\nMessage: {str(e)}")
        return (False, f"Error:\n{str(e)}")